"""Projects API routes."""

import json
import logging
from collections import defaultdict
from datetime import datetime
from typing import Annotated

import yaml
from fastapi import APIRouter, Depends, HTTPException, Query, status
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from induform.api.auth.dependencies import get_current_user
from induform.api.projects.schemas import (
    ComparisonResult,
    CsvImportResult,
    GrantAccessRequest,
    ImportYamlRequest,
    ProjectAccessInfo,
    ProjectCreate,
    ProjectDetail,
    ProjectSummary,
    ProjectUpdate,
)
from induform.db import ActivityLog, AssetDB, ProjectDB, User, Vulnerability, ZoneDB, get_db
from induform.db.repositories import ProjectRepository
from induform.engine.attack_path import AttackPathAnalysis, analyze_attack_paths
from induform.engine.gap_analysis import GapAnalysisReport, analyze_gaps
from induform.engine.policy import PolicySeverity, evaluate_policies
from induform.engine.risk import VulnInfo, assess_risk
from induform.models.project import Project
from induform.security.permissions import (
    Permission,
    check_project_permission,
    get_user_permission,
)

logger = logging.getLogger(__name__)


def _parse_compliance_standards(project_db: ProjectDB) -> list[str]:
    """Parse compliance_standards JSON from a DB project row."""
    if project_db.compliance_standards:
        try:
            return json.loads(project_db.compliance_standards)
        except (json.JSONDecodeError, TypeError):
            pass
    return [project_db.standard or "IEC62443"]


def _parse_allowed_protocols(project_db: ProjectDB) -> list[str]:
    """Parse allowed_protocols JSON from a DB project row."""
    if project_db.allowed_protocols:
        try:
            return json.loads(project_db.allowed_protocols)
        except (json.JSONDecodeError, TypeError):
            pass
    return []


async def _load_vulnerability_data(db: AsyncSession, project_id: str) -> dict[str, list[VulnInfo]]:
    """Load vulnerability data grouped by zone_id for risk scoring."""
    result = await db.execute(
        select(Vulnerability, ZoneDB.zone_id)
        .join(AssetDB, Vulnerability.asset_db_id == AssetDB.id)
        .join(ZoneDB, AssetDB.zone_db_id == ZoneDB.id)
        .where(ZoneDB.project_id == project_id)
    )
    rows = result.all()

    vuln_data: dict[str, list[VulnInfo]] = defaultdict(list)
    for vuln, zone_id in rows:
        vuln_data[zone_id].append(
            VulnInfo(
                cve_id=vuln.cve_id,
                severity=vuln.severity,
                cvss_score=vuln.cvss_score,
                status=vuln.status,
            )
        )
    return dict(vuln_data)


router = APIRouter(prefix="/projects", tags=["Projects"])


@router.get("/", response_model=list[ProjectSummary])
async def list_projects(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    skip: int = 0,
    limit: int = 100,
    include_archived: bool = Query(False, description="Include archived projects"),
) -> list[ProjectSummary]:
    """List all projects accessible to the current user."""
    project_repo = ProjectRepository(db)
    projects = await project_repo.list_accessible(
        current_user.id, skip, limit, load_full=True, is_admin=current_user.is_admin
    )

    result = []
    for project_db in projects:
        # Filter out archived projects if not requested
        if not include_archived and getattr(project_db, "is_archived", False):
            continue

        permission = await get_user_permission(
            db, project_db.id, current_user.id, is_admin=current_user.is_admin
        )

        # Calculate risk score and compliance if project has zones
        risk_score = None
        risk_level = None
        compliance_score = None
        zone_types: dict[str, int] = {}
        asset_count = 0

        if project_db.zones:
            try:
                project = await project_repo.to_pydantic(project_db)

                # Risk assessment
                vuln_data = await _load_vulnerability_data(db, project_db.id)
                risk_assessment = assess_risk(project, vulnerability_data=vuln_data)
                risk_score = int(round(risk_assessment.overall_score))
                risk_level = risk_assessment.overall_level.value

                # Compliance score based on policy violations
                enabled_standards = project.project.compliance_standards or None
                violations = evaluate_policies(project, enabled_standards=enabled_standards)
                if violations:
                    # Deduct points based on severity
                    deduction = 0
                    for v in violations:
                        if v.severity == PolicySeverity.CRITICAL:
                            deduction += 25
                        elif v.severity == PolicySeverity.HIGH:
                            deduction += 15
                        elif v.severity == PolicySeverity.MEDIUM:
                            deduction += 8
                        else:
                            deduction += 3
                    compliance_score = max(0, 100 - deduction)
                else:
                    compliance_score = 100

                # Zone types breakdown
                for zone in project.zones:
                    zone_type = zone.type.value if hasattr(zone.type, "value") else str(zone.type)
                    zone_types[zone_type] = zone_types.get(zone_type, 0) + 1

                # Asset count
                asset_count = sum(len(zone.assets) for zone in project.zones)
            except Exception:
                # If calculation fails, leave as None
                pass

        result.append(
            ProjectSummary(
                id=project_db.id,
                name=project_db.name,
                description=project_db.description,
                standard=project_db.standard,
                compliance_standards=_parse_compliance_standards(project_db),
                allowed_protocols=_parse_allowed_protocols(project_db),
                owner_id=project_db.owner_id,
                owner_username=project_db.owner.username if project_db.owner else None,
                created_at=project_db.created_at,
                updated_at=project_db.updated_at,
                zone_count=len(project_db.zones) if project_db.zones else 0,
                conduit_count=len(project_db.conduits) if project_db.conduits else 0,
                asset_count=asset_count,
                permission=permission.value if permission else "none",
                risk_score=risk_score,
                risk_level=risk_level,
                compliance_score=compliance_score,
                zone_types=zone_types if zone_types else None,
                is_archived=getattr(project_db, "is_archived", False),
                archived_at=getattr(project_db, "archived_at", None),
            )
        )

    return result


@router.post("/", response_model=ProjectSummary, status_code=status.HTTP_201_CREATED)
async def create_project(
    project_data: ProjectCreate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ProjectSummary:
    """Create a new project."""
    project_repo = ProjectRepository(db)

    project_db = await project_repo.create(
        name=project_data.name,
        owner_id=current_user.id,
        description=project_data.description,
        standard=project_data.standard,
        compliance_standards=project_data.compliance_standards,
        allowed_protocols=project_data.allowed_protocols,
    )

    # Log activity
    log = ActivityLog(
        project_id=project_db.id,
        user_id=current_user.id,
        action="created",
        entity_type="project",
        entity_id=project_db.id,
        entity_name=project_db.name,
    )
    db.add(log)
    await db.flush()

    return ProjectSummary(
        id=project_db.id,
        name=project_db.name,
        description=project_db.description,
        standard=project_db.standard,
        compliance_standards=_parse_compliance_standards(project_db),
        allowed_protocols=_parse_allowed_protocols(project_db),
        owner_id=project_db.owner_id,
        owner_username=current_user.username,
        created_at=project_db.created_at,
        updated_at=project_db.updated_at,
        zone_count=0,
        conduit_count=0,
        permission="owner",
        risk_score=None,
        risk_level=None,
    )


@router.get("/{project_id}", response_model=ProjectDetail)
async def get_project(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ProjectDetail:
    """Get a project by ID with full data."""
    project_repo = ProjectRepository(db)

    # Check permission
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Convert to Pydantic model
    project = await project_repo.to_pydantic(project_db)
    permission = await get_user_permission(
        db, project_id, current_user.id, is_admin=current_user.is_admin
    )

    return ProjectDetail(
        id=project_db.id,
        name=project_db.name,
        description=project_db.description,
        standard=project_db.standard,
        compliance_standards=_parse_compliance_standards(project_db),
        allowed_protocols=_parse_allowed_protocols(project_db),
        version=project_db.version,
        owner_id=project_db.owner_id,
        owner_username=project_db.owner.username if project_db.owner else None,
        created_at=project_db.created_at,
        updated_at=project_db.updated_at,
        permission=permission.value if permission else "none",
        project=project.model_dump(mode="json"),
    )


@router.put("/{project_id}", response_model=ProjectDetail)
async def update_project(
    project_id: str,
    project_data: Project,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ProjectDetail:
    """Update a project's data (zones, conduits, etc.)."""
    project_repo = ProjectRepository(db)

    # Check permission
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.EDITOR, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to edit this project",
        )

    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Capture old state for activity logging
    old_zone_ids = {z.zone_id for z in project_db.zones}
    old_conduit_ids = {c.conduit_id for c in project_db.conduits}
    old_asset_count = sum(len(z.assets) for z in project_db.zones)

    # Update project data
    await project_repo.from_pydantic(project_data, project_db)

    # Log structural changes (skip position-only updates)
    new_zone_ids = {z.id for z in project_data.zones}
    new_conduit_ids = {c.id for c in project_data.conduits}
    new_asset_count = sum(len(z.assets) for z in project_data.zones)

    added_zones = new_zone_ids - old_zone_ids
    removed_zones = old_zone_ids - new_zone_ids
    added_conduits = new_conduit_ids - old_conduit_ids
    removed_conduits = old_conduit_ids - new_conduit_ids
    asset_diff = new_asset_count - old_asset_count

    for zid in added_zones:
        zone = next((z for z in project_data.zones if z.id == zid), None)
        log = ActivityLog(
            project_id=project_id,
            user_id=current_user.id,
            action="zone_added",
            entity_type="zone",
            entity_id=zid,
            entity_name=zone.name if zone else zid,
        )
        db.add(log)

    for zid in removed_zones:
        log = ActivityLog(
            project_id=project_id,
            user_id=current_user.id,
            action="zone_deleted",
            entity_type="zone",
            entity_id=zid,
            entity_name=zid,
        )
        db.add(log)

    for cid in added_conduits:
        conduit = next((c for c in project_data.conduits if c.id == cid), None)
        log = ActivityLog(
            project_id=project_id,
            user_id=current_user.id,
            action="conduit_added",
            entity_type="conduit",
            entity_id=cid,
            entity_name=conduit.name or cid if conduit else cid,
        )
        db.add(log)

    for cid in removed_conduits:
        log = ActivityLog(
            project_id=project_id,
            user_id=current_user.id,
            action="conduit_deleted",
            entity_type="conduit",
            entity_id=cid,
            entity_name=cid,
        )
        db.add(log)

    if asset_diff > 0:
        log = ActivityLog(
            project_id=project_id,
            user_id=current_user.id,
            action="asset_added",
            entity_type="asset",
            entity_name=f"{asset_diff} asset(s) added",
        )
        db.add(log)
    elif asset_diff < 0:
        log = ActivityLog(
            project_id=project_id,
            user_id=current_user.id,
            action="asset_deleted",
            entity_type="asset",
            entity_name=f"{abs(asset_diff)} asset(s) removed",
        )
        db.add(log)

    # Notify collaborators of structural changes
    structural_changes = added_zones | removed_zones | added_conduits | removed_conduits
    if structural_changes or asset_diff != 0:
        try:
            from induform.api.notifications.routes import create_notification

            access_list = await project_repo.list_access(project_id)
            collaborator_ids = {
                a.user_id for a in access_list if a.user_id and a.user_id != current_user.id
            }
            if project_db.owner_id != current_user.id:
                collaborator_ids.add(project_db.owner_id)
            parts = []
            if added_zones:
                parts.append(f"{len(added_zones)} zone(s) added")
            if removed_zones:
                parts.append(f"{len(removed_zones)} zone(s) removed")
            if added_conduits:
                parts.append(f"{len(added_conduits)} conduit(s) added")
            if removed_conduits:
                parts.append(f"{len(removed_conduits)} conduit(s) removed")
            if asset_diff > 0:
                parts.append(f"{asset_diff} asset(s) added")
            elif asset_diff < 0:
                parts.append(f"{abs(asset_diff)} asset(s) removed")
            change_summary = ", ".join(parts)
            for uid in collaborator_ids:
                await create_notification(
                    db,
                    user_id=uid,
                    type="project_update",
                    title=f"Project updated: {project_db.name}",
                    message=f"{current_user.username} made changes: {change_summary}",
                    link=f"/projects/{project_id}",
                    project_id=project_id,
                    actor_id=current_user.id,
                )
        except Exception as e:
            logger.warning("Failed to create notifications for project %s: %s", project_id, e)

    # Create automatic version snapshot
    try:
        from induform.api.versions.routes import create_auto_version

        await create_auto_version(db, project_id, current_user.id, "Auto-save")
    except Exception as e:
        logger.warning("Failed to create auto-version for project %s: %s", project_id, e)

    # Reload and return
    project_db = await project_repo.get_by_id(project_id)
    project = await project_repo.to_pydantic(project_db)
    permission = await get_user_permission(
        db, project_id, current_user.id, is_admin=current_user.is_admin
    )

    # Record metrics snapshot (throttled to max 1 per 5 min per project)
    try:
        await _record_metrics_snapshot(db, project_id, project)
    except Exception as e:
        logger.warning("Failed to record metrics snapshot for project %s: %s", project_id, e)

    return ProjectDetail(
        id=project_db.id,
        name=project_db.name,
        description=project_db.description,
        standard=project_db.standard,
        compliance_standards=_parse_compliance_standards(project_db),
        allowed_protocols=_parse_allowed_protocols(project_db),
        version=project_db.version,
        owner_id=project_db.owner_id,
        owner_username=project_db.owner.username if project_db.owner else None,
        created_at=project_db.created_at,
        updated_at=project_db.updated_at,
        permission=permission.value if permission else "none",
        project=project.model_dump(mode="json"),
    )


@router.patch("/{project_id}", response_model=ProjectSummary)
async def update_project_metadata(
    project_id: str,
    update_data: ProjectUpdate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ProjectSummary:
    """Update a project's metadata (name, description)."""
    project_repo = ProjectRepository(db)

    # Check permission (owner or editor)
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.EDITOR, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to edit this project",
        )

    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Update metadata
    update_fields = update_data.model_dump(exclude_unset=True)
    if update_fields:
        await project_repo.update(project_db, **update_fields)

        # Log metadata update
        changed_keys = list(update_fields.keys())
        log = ActivityLog(
            project_id=project_id,
            user_id=current_user.id,
            action="updated",
            entity_type="project",
            entity_id=project_id,
            entity_name=project_db.name,
            details=json.dumps({"fields": changed_keys}),
        )
        db.add(log)

    permission = await get_user_permission(
        db, project_id, current_user.id, is_admin=current_user.is_admin
    )

    # Calculate risk score if project has zones
    risk_score = None
    risk_level = None
    if project_db.zones:
        try:
            project = await project_repo.to_pydantic(project_db)
            vuln_data = await _load_vulnerability_data(db, project_id)
            risk_assessment = assess_risk(project, vulnerability_data=vuln_data)
            risk_score = int(round(risk_assessment.overall_score))
            risk_level = risk_assessment.overall_level.value
        except Exception:
            pass

    return ProjectSummary(
        id=project_db.id,
        name=project_db.name,
        description=project_db.description,
        standard=project_db.standard,
        compliance_standards=_parse_compliance_standards(project_db),
        allowed_protocols=_parse_allowed_protocols(project_db),
        owner_id=project_db.owner_id,
        owner_username=project_db.owner.username if project_db.owner else None,
        created_at=project_db.created_at,
        updated_at=project_db.updated_at,
        zone_count=len(project_db.zones) if project_db.zones else 0,
        conduit_count=len(project_db.conduits) if project_db.conduits else 0,
        permission=permission.value if permission else "none",
        risk_score=risk_score,
        risk_level=risk_level,
    )


@router.delete("/{project_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_project(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    """Delete a project. Only the owner can delete."""
    project_repo = ProjectRepository(db)

    project_db = await project_repo.get_by_id(project_id, load_relations=False)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Check permission (owner or editor)
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.EDITOR, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to delete this project",
        )

    # Log before deletion
    log = ActivityLog(
        project_id=project_id,
        user_id=current_user.id,
        action="deleted",
        entity_type="project",
        entity_id=project_id,
        entity_name=project_db.name,
    )
    db.add(log)
    await db.flush()

    await project_repo.delete(project_db)


# Duplication endpoint


@router.post(
    "/{project_id}/duplicate", response_model=ProjectSummary, status_code=status.HTTP_201_CREATED
)
async def duplicate_project(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    name: str | None = None,
) -> ProjectSummary:
    """Duplicate a project. Requires at least viewer permission on the source."""
    project_repo = ProjectRepository(db)

    # Check permission (viewer can duplicate to create their own copy)
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Duplicate the project
    new_project = await project_repo.duplicate(
        source_project_id=project_id,
        new_owner_id=current_user.id,
        new_name=name,
    )

    if not new_project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Calculate risk score for the duplicated project
    risk_score = None
    risk_level = None
    if new_project.zones:
        try:
            project = await project_repo.to_pydantic(new_project)
            vuln_data = await _load_vulnerability_data(db, new_project.id)
            risk_assessment = assess_risk(project, vulnerability_data=vuln_data)
            risk_score = int(round(risk_assessment.overall_score))
            risk_level = risk_assessment.overall_level.value
        except Exception:
            pass

    return ProjectSummary(
        id=new_project.id,
        name=new_project.name,
        description=new_project.description,
        standard=new_project.standard,
        compliance_standards=_parse_compliance_standards(new_project),
        allowed_protocols=_parse_allowed_protocols(new_project),
        owner_id=new_project.owner_id,
        owner_username=current_user.username,
        created_at=new_project.created_at,
        updated_at=new_project.updated_at,
        zone_count=len(new_project.zones) if new_project.zones else 0,
        conduit_count=len(new_project.conduits) if new_project.conduits else 0,
        permission="owner",
        risk_score=risk_score,
        risk_level=risk_level,
    )


# Access control endpoints


@router.get("/{project_id}/access", response_model=list[ProjectAccessInfo])
async def list_project_access(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> list[ProjectAccessInfo]:
    """List all access grants for a project."""
    project_repo = ProjectRepository(db)

    # Check permission (owner or editor)
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.EDITOR, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to view access list",
        )

    project_db = await project_repo.get_by_id(project_id, load_relations=False)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    access_list = await project_repo.list_access(project_id)

    return [
        ProjectAccessInfo(
            id=access.id,
            user_id=access.user_id,
            user_email=access.user.email if access.user else None,
            user_username=access.user.username if access.user else None,
            team_id=access.team_id,
            team_name=access.team.name if access.team else None,
            permission=access.permission,
            granted_by=access.granted_by,
            granted_at=access.granted_at,
        )
        for access in access_list
    ]


@router.post(
    "/{project_id}/access", response_model=ProjectAccessInfo, status_code=status.HTTP_201_CREATED
)
async def grant_project_access(
    project_id: str,
    access_data: GrantAccessRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ProjectAccessInfo:
    """Grant access to a project."""
    project_repo = ProjectRepository(db)

    # Check permission (owner or editor)
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.EDITOR, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to share this project",
        )

    project_db = await project_repo.get_by_id(project_id, load_relations=False)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    if not access_data.user_id and not access_data.team_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Either user_id or team_id must be provided",
        )

    access = await project_repo.grant_access(
        project_id=project_id,
        granted_by=current_user.id,
        user_id=access_data.user_id,
        team_id=access_data.team_id,
        permission=access_data.permission,
    )

    # Log sharing
    log = ActivityLog(
        project_id=project_id,
        user_id=current_user.id,
        action="shared",
        entity_type="access",
        entity_name=f"Shared as {access_data.permission}",
        details=json.dumps(
            {
                "user_id": access_data.user_id,
                "team_id": access_data.team_id,
                "permission": access_data.permission,
            }
        ),
    )
    db.add(log)

    # Notify the user being granted access
    if access_data.user_id:
        try:
            from induform.api.notifications.routes import create_notification

            await create_notification(
                db,
                user_id=access_data.user_id,
                type="share",
                title=f"Project shared with you: {project_db.name}",
                message=(
                    f"{current_user.username} shared a project with you"
                    f" ({access_data.permission} access)"
                ),
                link=f"/projects/{project_id}",
                project_id=project_id,
                actor_id=current_user.id,
            )
        except Exception as e:
            logger.warning("Failed to create share notification: %s", e)

    return ProjectAccessInfo(
        id=access.id,
        user_id=access.user_id,
        user_email=None,
        user_username=None,
        team_id=access.team_id,
        team_name=None,
        permission=access.permission,
        granted_by=access.granted_by,
        granted_at=access.granted_at,
    )


@router.delete("/{project_id}/access/{access_id}", status_code=status.HTTP_204_NO_CONTENT)
async def revoke_project_access(
    project_id: str,
    access_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    """Revoke access to a project."""
    project_repo = ProjectRepository(db)

    # Check permission (owner or editor)
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.EDITOR, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to manage access",
        )

    project_db = await project_repo.get_by_id(project_id, load_relations=False)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Look up the access record to get the affected user before revoking
    from sqlalchemy import select as sa_select

    from induform.db.models import ProjectAccess

    access_query = sa_select(ProjectAccess).where(ProjectAccess.id == access_id)
    access_result = await db.execute(access_query)
    access_record = access_result.scalar_one_or_none()
    revoked_user_id = access_record.user_id if access_record else None

    success = await project_repo.revoke_access(access_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Access grant not found",
        )

    # Log access revocation
    log = ActivityLog(
        project_id=project_id,
        user_id=current_user.id,
        action="access_revoked",
        entity_type="access",
        entity_name="Access revoked",
    )
    db.add(log)

    # Notify the user whose access was revoked
    if revoked_user_id and revoked_user_id != current_user.id:
        try:
            from induform.api.notifications.routes import create_notification

            await create_notification(
                db,
                user_id=revoked_user_id,
                type="access_revoked",
                title=f"Access revoked: {project_db.name}",
                message=f"{current_user.username} removed your access to this project",
                project_id=project_id,
                actor_id=current_user.id,
            )
        except Exception as e:
            logger.warning("Failed to create revoke notification: %s", e)


# YAML import/export endpoints


@router.post("/{project_id}/export/yaml")
async def export_project_yaml(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, str]:
    """Export a project as YAML."""
    project_repo = ProjectRepository(db)

    # Check permission
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Convert to Pydantic and then to YAML
    project = await project_repo.to_pydantic(project_db)
    yaml_content = yaml.dump(
        project.model_dump(mode="json", exclude_none=True),
        default_flow_style=False,
        sort_keys=False,
    )

    return {"yaml": yaml_content, "filename": f"{project_db.name.lower().replace(' ', '_')}.yaml"}


@router.post("/import/yaml", response_model=ProjectSummary, status_code=status.HTTP_201_CREATED)
async def import_project_yaml(
    import_data: ImportYamlRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ProjectSummary:
    """Import a project from YAML content."""
    project_repo = ProjectRepository(db)

    try:
        data = yaml.safe_load(import_data.yaml_content)
        project = Project.model_validate(data)
    except yaml.YAMLError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid YAML: {e}",
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid project data: {e}",
        )

    # Override name if provided
    if import_data.name:
        project.project.name = import_data.name

    # Create project from Pydantic model
    project_db = await project_repo.create_from_pydantic(project, current_user.id)

    # Calculate risk score for the imported project
    risk_score = None
    risk_level = None
    if project_db.zones:
        try:
            vuln_data = await _load_vulnerability_data(db, project_db.id)
            risk_assessment = assess_risk(project, vulnerability_data=vuln_data)
            risk_score = int(round(risk_assessment.overall_score))
            risk_level = risk_assessment.overall_level.value
        except Exception:
            pass

    return ProjectSummary(
        id=project_db.id,
        name=project_db.name,
        description=project_db.description,
        standard=project_db.standard,
        compliance_standards=_parse_compliance_standards(project_db),
        allowed_protocols=_parse_allowed_protocols(project_db),
        owner_id=project_db.owner_id,
        owner_username=current_user.username,
        created_at=project_db.created_at,
        updated_at=project_db.updated_at,
        zone_count=len(project_db.zones) if project_db.zones else 0,
        conduit_count=len(project_db.conduits) if project_db.conduits else 0,
        permission="owner",
        risk_score=risk_score,
        risk_level=risk_level,
    )


# Archive/Restore endpoints


@router.post("/{project_id}/archive")
async def archive_project(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    """Archive a project."""
    project_repo = ProjectRepository(db)

    # Check permission (owner or editor)
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.EDITOR, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to archive this project",
        )

    project_db = await project_repo.get_by_id(project_id, load_relations=False)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    if project_db.is_archived:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Project is already archived",
        )

    project_db.is_archived = True
    project_db.archived_at = datetime.utcnow()
    await db.flush()

    # Log activity
    log = ActivityLog(
        project_id=project_id,
        user_id=current_user.id,
        action="archived",
        entity_type="project",
        entity_id=project_id,
        entity_name=project_db.name,
    )
    db.add(log)

    return {"message": "Project archived successfully"}


@router.post("/{project_id}/restore")
async def restore_project(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    """Restore an archived project."""
    project_repo = ProjectRepository(db)

    # Check permission (owner or editor)
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.EDITOR, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to restore this project",
        )

    project_db = await project_repo.get_by_id(project_id, load_relations=False)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    if not project_db.is_archived:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Project is not archived",
        )

    project_db.is_archived = False
    project_db.archived_at = None
    await db.flush()

    # Log activity
    log = ActivityLog(
        project_id=project_id,
        user_id=current_user.id,
        action="restored",
        entity_type="project",
        entity_id=project_id,
        entity_name=project_db.name,
    )
    db.add(log)

    return {"message": "Project restored successfully"}


# Bulk operations


class BulkOperationRequest(BaseModel):
    """Bulk operation request."""

    project_ids: list[str]
    operation: str  # "archive", "restore", "delete", "export"


class BulkOperationResult(BaseModel):
    """Bulk operation result."""

    success: list[str]
    failed: list[dict]


@router.post("/bulk", response_model=BulkOperationResult)
async def bulk_operation(
    request: BulkOperationRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> BulkOperationResult:
    """Perform bulk operations on projects."""
    project_repo = ProjectRepository(db)

    success = []
    failed = []

    for project_id in request.project_ids:
        try:
            project_db = await project_repo.get_by_id(project_id, load_relations=False)
            if not project_db:
                failed.append({"id": project_id, "error": "Project not found"})
                continue

            has_access = await check_project_permission(
                db, project_id, current_user.id, Permission.EDITOR, is_admin=current_user.is_admin
            )
            if not has_access:
                failed.append({"id": project_id, "error": "No permission"})
                continue

            if request.operation == "archive":
                if project_db.is_archived:
                    failed.append({"id": project_id, "error": "Already archived"})
                    continue
                project_db.is_archived = True
                project_db.archived_at = datetime.utcnow()
                success.append(project_id)

            elif request.operation == "restore":
                if not project_db.is_archived:
                    failed.append({"id": project_id, "error": "Not archived"})
                    continue
                project_db.is_archived = False
                project_db.archived_at = None
                success.append(project_id)

            elif request.operation == "delete":
                await project_repo.delete(project_db)
                success.append(project_id)

            else:
                failed.append(
                    {"id": project_id, "error": f"Unknown operation: {request.operation}"}
                )

        except Exception as e:
            failed.append({"id": project_id, "error": str(e)})

    return BulkOperationResult(success=success, failed=failed)


# JSON export endpoint
@router.post("/{project_id}/export/json")
async def export_project_json(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    """Export a project as JSON."""
    project_repo = ProjectRepository(db)

    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project = await project_repo.to_pydantic(project_db)
    return {
        "json": project.model_dump(mode="json", exclude_none=True),
        "filename": f"{project_db.name.lower().replace(' ', '_')}.json",
    }


# CSV Import endpoint
class CsvImportRequest(BaseModel):
    """Request to import assets from CSV."""

    csv_content: str
    zone_id: str


@router.post("/{project_id}/import/csv", response_model=CsvImportResult)
async def import_assets_csv(
    project_id: str,
    request: CsvImportRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> CsvImportResult:
    """
    Import assets from CSV into a specific zone.

    CSV columns (header required):
    - id: Unique asset ID (required)
    - name: Asset name (required)
    - type: Asset type (plc, hmi, scada, etc.)
    - ip_address: IP address
    - mac_address: MAC address
    - vendor: Equipment vendor
    - model: Equipment model
    - criticality: 1-5 criticality level
    - description: Additional description
    """
    import csv
    import io

    from induform.models.asset import Asset, AssetType

    project_repo = ProjectRepository(db)

    # Check permission
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.EDITOR, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to edit this project",
        )

    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Get current project state
    project = await project_repo.to_pydantic(project_db)

    # Find the target zone
    target_zone = None
    for zone in project.zones:
        if zone.id == request.zone_id:
            target_zone = zone
            break

    if not target_zone:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Zone '{request.zone_id}' not found in project",
        )

    # Parse CSV
    imported = 0
    skipped = 0
    errors = []

    reader = csv.DictReader(io.StringIO(request.csv_content))
    existing_ids = {a.id for a in target_zone.assets}

    for row_num, row in enumerate(reader, start=2):
        try:
            asset_id = row.get("id", "").strip()
            name = row.get("name", "").strip()

            if not asset_id or not name:
                errors.append({"row": str(row_num), "error": "Missing required field: id or name"})
                skipped += 1
                continue

            if asset_id in existing_ids:
                errors.append(
                    {"row": str(row_num), "error": f"Asset ID '{asset_id}' already exists"}
                )
                skipped += 1
                continue

            # Parse asset type
            asset_type_str = row.get("type", "other").strip().lower()
            try:
                asset_type = AssetType(asset_type_str)
            except ValueError:
                asset_type = AssetType.OTHER

            # Parse criticality
            try:
                criticality = int(row.get("criticality", "3") or "3")
                criticality = max(1, min(5, criticality))
            except ValueError:
                criticality = 3

            # Parse vlan as integer
            vlan_val = None
            vlan_str = row.get("vlan", "").strip()
            if vlan_str:
                try:
                    vlan_val = int(vlan_str)
                except ValueError:
                    pass

            asset = Asset(
                id=asset_id,
                name=name,
                type=asset_type,
                ip_address=row.get("ip_address", "").strip() or None,
                mac_address=row.get("mac_address", "").strip() or None,
                vendor=row.get("vendor", "").strip() or None,
                model=row.get("model", "").strip() or None,
                criticality=criticality,
                description=row.get("description", "").strip() or None,
                firmware_version=row.get("firmware_version", "").strip() or None,
                os_name=row.get("os_name", "").strip() or None,
                os_version=row.get("os_version", "").strip() or None,
                software=row.get("software", "").strip() or None,
                cpe=row.get("cpe", "").strip() or None,
                subnet=row.get("subnet", "").strip() or None,
                gateway=row.get("gateway", "").strip() or None,
                vlan=vlan_val,
                dns=row.get("dns", "").strip() or None,
                open_ports=row.get("open_ports", "").strip() or None,
                protocols=row.get("protocols", "").strip() or None,
                purchase_date=row.get("purchase_date", "").strip() or None,
                end_of_life=row.get("end_of_life", "").strip() or None,
                warranty_expiry=row.get("warranty_expiry", "").strip() or None,
                last_patched=row.get("last_patched", "").strip() or None,
                patch_level=row.get("patch_level", "").strip() or None,
                location=row.get("location", "").strip() or None,
            )

            target_zone.assets.append(asset)
            existing_ids.add(asset_id)
            imported += 1

        except Exception as e:
            errors.append({"row": str(row_num), "error": str(e)})
            skipped += 1

    # Save updated project
    if imported > 0:
        await project_repo.from_pydantic(project, project_db)

        # Log activity
        log = ActivityLog(
            project_id=project_id,
            user_id=current_user.id,
            action="asset_added",
            entity_type="zone",
            entity_id=request.zone_id,
            entity_name=f"{imported} assets imported",
            details=json.dumps({"zone_id": request.zone_id, "count": imported}),
        )
        db.add(log)

    return CsvImportResult(
        imported=imported,
        skipped=skipped,
        errors=errors[:10],  # Limit errors to first 10
    )


# CSV Asset Export endpoint
@router.get("/{project_id}/export/assets-csv")
async def export_assets_csv(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Export all project assets as CSV with all fields including zone info."""
    import csv
    import io

    from starlette.responses import StreamingResponse

    project_repo = ProjectRepository(db)

    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project = await project_repo.to_pydantic(project_db)

    headers = [
        "zone_id",
        "zone_name",
        "id",
        "name",
        "type",
        "ip_address",
        "mac_address",
        "vendor",
        "model",
        "firmware_version",
        "criticality",
        "description",
        "os_name",
        "os_version",
        "software",
        "cpe",
        "subnet",
        "gateway",
        "vlan",
        "dns",
        "open_ports",
        "protocols",
        "purchase_date",
        "end_of_life",
        "warranty_expiry",
        "last_patched",
        "patch_level",
        "location",
    ]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for zone in project.zones:
        for asset in zone.assets:
            writer.writerow(
                [
                    zone.id,
                    zone.name,
                    asset.id,
                    asset.name,
                    asset.type,
                    asset.ip_address or "",
                    asset.mac_address or "",
                    asset.vendor or "",
                    asset.model or "",
                    asset.firmware_version or "",
                    asset.criticality if asset.criticality is not None else "",
                    asset.description or "",
                    asset.os_name or "",
                    asset.os_version or "",
                    asset.software or "",
                    asset.cpe or "",
                    asset.subnet or "",
                    asset.gateway or "",
                    asset.vlan if asset.vlan is not None else "",
                    asset.dns or "",
                    asset.open_ports or "",
                    asset.protocols or "",
                    asset.purchase_date or "",
                    asset.end_of_life or "",
                    asset.warranty_expiry or "",
                    asset.last_patched or "",
                    asset.patch_level or "",
                    asset.location or "",
                ]
            )

    csv_content = output.getvalue()
    filename = f"{project_db.name.lower().replace(' ', '_')}_assets.csv"

    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{project_id}/export/assets-csv-template")
async def export_assets_csv_template(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Download a blank CSV template for asset import."""
    import csv
    import io

    from starlette.responses import StreamingResponse

    # Just verify project access
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    headers = [
        "zone_id",
        "id",
        "name",
        "type",
        "ip_address",
        "mac_address",
        "vendor",
        "model",
        "firmware_version",
        "criticality",
        "description",
        "os_name",
        "os_version",
        "software",
        "cpe",
        "subnet",
        "gateway",
        "vlan",
        "dns",
        "open_ports",
        "protocols",
        "purchase_date",
        "end_of_life",
        "warranty_expiry",
        "last_patched",
        "patch_level",
        "location",
    ]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    # Add one example row
    writer.writerow(
        [
            "cell_01",
            "plc_01",
            "Main PLC",
            "plc",
            "10.10.1.10",
            "00:1A:2B:3C:4D:5E",
            "Siemens",
            "S7-1500",
            "4.5.2",
            "4",
            "Primary production controller",
            "Linux",
            "4.19",
            "Step 7",
            "cpe:2.3:h:siemens:s7-1500:-:*:*:*:*:*:*:*",
            "10.10.1.0/24",
            "10.10.1.1",
            "100",
            "10.10.1.1",
            "102,502",
            "S7,Modbus",
            "2023-01-15",
            "2030-12-31",
            "2028-01-15",
            "2025-06-01",
            "SP3",
            "Building A, Rack 2",
        ]
    )

    csv_content = output.getvalue()
    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="asset_import_template.csv"'},
    )


# Excel export endpoint
@router.post("/{project_id}/export/excel")
async def export_project_excel(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    """
    Export a project to Excel format.

    Creates a workbook with sheets for:
    - Summary: Project metadata
    - Zones: All zones with their properties
    - Assets: All assets organized by zone
    - Conduits: All conduits with flow information
    """
    import base64
    import io

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export requires the 'openpyxl' package. "
            "Install with: pip install openpyxl",
        )

    project_repo = ProjectRepository(db)

    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project = await project_repo.to_pydantic(project_db)

    # Create workbook
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, row=1, cols=None):
        if cols is None:
            cols = range(1, ws.max_column + 1)
        for col in cols:
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_data = [
        ["Project Name", project_db.name],
        ["Description", project_db.description or ""],
        [
            "Standard",
            ", ".join(project.project.compliance_standards) if project.project else "IEC62443",
        ],
        ["Version", project.version],
        ["Total Zones", len(project.zones)],
        ["Total Conduits", len(project.conduits)],
        ["Total Assets", sum(len(z.assets) for z in project.zones)],
    ]
    for row in summary_data:
        ws_summary.append(row)
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 50

    # Zones sheet
    ws_zones = wb.create_sheet("Zones")
    zone_headers = [
        "ID",
        "Name",
        "Type",
        "Security Level Target",
        "Parent Zone",
        "Description",
        "Assets Count",
    ]
    ws_zones.append(zone_headers)
    style_header(ws_zones)

    for zone in project.zones:
        ws_zones.append(
            [
                zone.id,
                zone.name,
                zone.type,
                zone.security_level_target,
                zone.parent_zone or "",
                zone.description or "",
                len(zone.assets),
            ]
        )

    for col_num, width in enumerate([15, 25, 15, 20, 15, 40, 15], 1):
        ws_zones.column_dimensions[chr(64 + col_num)].width = width

    # Assets sheet
    ws_assets = wb.create_sheet("Assets")
    asset_headers = [
        "Zone ID",
        "Asset ID",
        "Name",
        "Type",
        "IP Address",
        "Vendor",
        "Model",
        "Criticality",
        "Description",
    ]
    ws_assets.append(asset_headers)
    style_header(ws_assets)

    for zone in project.zones:
        for asset in zone.assets:
            ws_assets.append(
                [
                    zone.id,
                    asset.id,
                    asset.name,
                    asset.type,
                    asset.ip_address or "",
                    asset.vendor or "",
                    asset.model or "",
                    asset.criticality,
                    asset.description or "",
                ]
            )

    for col_num, width in enumerate([15, 15, 25, 20, 15, 15, 15, 12, 40], 1):
        ws_assets.column_dimensions[chr(64 + col_num)].width = width

    # Conduits sheet
    ws_conduits = wb.create_sheet("Conduits")
    conduit_headers = [
        "ID",
        "Name",
        "From Zone",
        "To Zone",
        "Security Level Required",
        "Requires Inspection",
        "Protocols",
    ]
    ws_conduits.append(conduit_headers)
    style_header(ws_conduits)

    for conduit in project.conduits:
        protocols = ", ".join(
            [f"{f.protocol}:{f.port}" if f.port else f.protocol for f in conduit.flows]
        )
        ws_conduits.append(
            [
                conduit.id,
                conduit.name or "",
                conduit.from_zone,
                conduit.to_zone,
                conduit.security_level_required or "",
                "Yes" if conduit.requires_inspection else "No",
                protocols,
            ]
        )

    for col_num, width in enumerate([15, 25, 15, 15, 22, 18, 40], 1):
        ws_conduits.column_dimensions[chr(64 + col_num)].width = width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Return as base64 encoded string
    excel_data = base64.b64encode(output.getvalue()).decode("utf-8")
    filename = f"{project_db.name.lower().replace(' ', '_')}.xlsx"

    return {"excel_base64": excel_data, "filename": filename}


def _draw_risk_matrix(
    risk_result,
    zones: list,
):
    """Render a 5×5 risk matrix heatmap as a ReportLab Drawing."""
    try:
        from reportlab.graphics.shapes import Drawing, Rect, String
        from reportlab.lib import colors as gcolors
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="PDF export requires the 'reportlab' package. "
            "Install with: pip install reportlab",
        )

    cell_w, cell_h = 52, 40
    margin_left, margin_bottom = 80, 40
    width = margin_left + 5 * cell_w + 20
    height = margin_bottom + 5 * cell_h + 50

    d = Drawing(width, height)

    # Risk level color for each cell (row=impact 0-4, col=likelihood 0-4)
    # Higher row+col = higher risk
    cell_colors = [
        ["#22c55e", "#22c55e", "#eab308", "#eab308", "#f97316"],  # row 0 (Negligible)
        ["#22c55e", "#eab308", "#eab308", "#f97316", "#f97316"],  # row 1 (Minor)
        ["#eab308", "#eab308", "#f97316", "#f97316", "#ef4444"],  # row 2 (Moderate)
        ["#eab308", "#f97316", "#f97316", "#ef4444", "#ef4444"],  # row 3 (Major)
        ["#f97316", "#f97316", "#ef4444", "#ef4444", "#ef4444"],  # row 4 (Catastrophic)
    ]

    impact_labels = ["Negligible", "Minor", "Moderate", "Major", "Catastrophic"]
    likelihood_labels = ["Rare", "Unlikely", "Possible", "Likely", "Almost\nCertain"]

    # Draw cells
    for row in range(5):
        for col in range(5):
            x = margin_left + col * cell_w
            y = margin_bottom + row * cell_h
            r = Rect(x, y, cell_w, cell_h)
            r.fillColor = gcolors.HexColor(cell_colors[row][col])
            r.strokeColor = gcolors.HexColor("#94a3b8")
            r.strokeWidth = 0.5
            d.add(r)

    # Place zone abbreviations in cells based on risk score and SL-T
    if risk_result.zone_risks:
        for zone_id, zr in risk_result.zone_risks.items():
            # Find the zone to get its SL-T
            zone = next((z for z in zones if z.id == zone_id), None)
            if not zone:
                continue
            # Impact: SL-T 1=row0 ... SL-T 4=row3, default row2
            impact_row = min(max(zone.security_level_target - 1, 0), 4)
            # Likelihood: risk score 0-20=col0 ... 80-100=col4
            likelihood_col = min(int(zr.score / 20), 4)
            x = margin_left + likelihood_col * cell_w + cell_w / 2
            y = margin_bottom + impact_row * cell_h + cell_h / 2
            label = zone.name[:6]
            s = String(x, y - 4, label, fontSize=7, textAnchor="middle")
            s.fillColor = gcolors.white
            d.add(s)

    # Y-axis labels (Impact)
    for i, label in enumerate(impact_labels):
        y = margin_bottom + i * cell_h + cell_h / 2
        s = String(margin_left - 5, y - 4, label, fontSize=7, textAnchor="end")
        s.fillColor = gcolors.HexColor("#334155")
        d.add(s)

    # X-axis labels (Likelihood)
    for i, label in enumerate(likelihood_labels):
        x = margin_left + i * cell_w + cell_w / 2
        # Handle multi-line by using first word only
        display = label.split("\n")[0]
        s = String(x, margin_bottom - 12, display, fontSize=7, textAnchor="middle")
        s.fillColor = gcolors.HexColor("#334155")
        d.add(s)

    # Axis titles
    s = String(
        margin_left + 5 * cell_w / 2,
        margin_bottom - 28,
        "Likelihood →",
        fontSize=8,
        textAnchor="middle",
    )
    s.fillColor = gcolors.HexColor("#1e40af")
    d.add(s)

    s = String(8, margin_bottom + 5 * cell_h / 2, "Impact →", fontSize=8, textAnchor="middle")
    s.fillColor = gcolors.HexColor("#1e40af")
    d.add(s)

    return d


def _draw_topology_diagram(
    zones: list,
    conduits: list,
):
    """Render a zone/conduit topology diagram as a ReportLab Drawing."""
    try:
        from reportlab.graphics.shapes import Drawing, Line, Rect, String
        from reportlab.lib import colors as gcolors
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="PDF export requires the 'reportlab' package. "
            "Install with: pip install reportlab",
        )

    # Layer order (Purdue-like, top to bottom)
    layer_order = ["enterprise", "dmz", "site", "area", "cell", "safety"]
    layer_y = {name: (5 - i) * 70 + 40 for i, name in enumerate(layer_order)}

    # Group zones by layer
    layers: dict[str, list] = {name: [] for name in layer_order}
    for zone in zones:
        zone_type = zone.type.value if hasattr(zone.type, "value") else str(zone.type)
        if zone_type in layers:
            layers[zone_type].append(zone)
        else:
            layers["area"].append(zone)  # fallback

    # Calculate dimensions
    max_per_layer = max((len(v) for v in layers.values() if v), default=1)
    box_w, box_h = 100, 35
    h_spacing = 120
    width = max(max_per_layer * h_spacing + 60, 500)
    height = 6 * 70 + 100

    d = Drawing(width, height)

    # Zone type colors
    type_colors = {
        "enterprise": "#3b82f6",
        "dmz": "#f59e0b",
        "site": "#8b5cf6",
        "area": "#06b6d4",
        "cell": "#10b981",
        "safety": "#ef4444",
    }

    # Draw zones and record positions
    zone_positions: dict[str, tuple[float, float]] = {}
    for layer_name, layer_zones in layers.items():
        if not layer_zones:
            continue
        y = layer_y[layer_name]
        total_width = len(layer_zones) * h_spacing
        start_x = (width - total_width) / 2 + h_spacing / 2 - box_w / 2

        # Layer label
        label_s = String(15, y + box_h / 2 - 4, layer_name.upper(), fontSize=6, textAnchor="start")
        label_s.fillColor = gcolors.HexColor("#94a3b8")
        d.add(label_s)

        for i, zone in enumerate(layer_zones):
            x = start_x + i * h_spacing
            center_x = x + box_w / 2
            center_y = y + box_h / 2
            zone_positions[zone.id] = (center_x, center_y)

            color = type_colors.get(
                zone.type.value if hasattr(zone.type, "value") else str(zone.type),
                "#64748b",
            )
            r = Rect(x, y, box_w, box_h, rx=4, ry=4)
            r.fillColor = gcolors.HexColor(color)
            r.strokeColor = gcolors.HexColor("#1e293b")
            r.strokeWidth = 0.8
            d.add(r)

            # Zone name (truncated)
            name = zone.name[:14]
            s = String(center_x, center_y, name, fontSize=7, textAnchor="middle")
            s.fillColor = gcolors.white
            d.add(s)

            # SL-T badge
            sl = String(
                center_x,
                y + 5,
                f"SL-{zone.security_level_target}",
                fontSize=6,
                textAnchor="middle",
            )
            sl.fillColor = gcolors.HexColor("#e2e8f0")
            d.add(sl)

    # Draw conduits as lines
    for conduit in conduits:
        from_pos = zone_positions.get(conduit.from_zone)
        to_pos = zone_positions.get(conduit.to_zone)
        if not from_pos or not to_pos:
            continue

        color = "#22c55e" if conduit.requires_inspection else "#ef4444"
        line = Line(from_pos[0], from_pos[1], to_pos[0], to_pos[1])
        line.strokeColor = gcolors.HexColor(color)
        line.strokeWidth = 1.2
        d.add(line)

    # Legend
    legend_y = 10
    legend_items = [
        ("#22c55e", "Inspected conduit"),
        ("#ef4444", "Uninspected conduit"),
    ]
    legend_x = 10
    for color, label in legend_items:
        line = Line(legend_x, legend_y, legend_x + 20, legend_y)
        line.strokeColor = gcolors.HexColor(color)
        line.strokeWidth = 2
        d.add(line)
        s = String(legend_x + 25, legend_y - 3, label, fontSize=6, textAnchor="start")
        s.fillColor = gcolors.HexColor("#334155")
        d.add(s)
        legend_x += 120

    return d


# PDF Report endpoint
@router.post("/{project_id}/export/pdf")
async def export_project_pdf(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    """
    Generate a PDF security report for the project.

    Creates a comprehensive PDF document with:
    - Table of contents
    - Executive summary
    - Zone inventory with security levels
    - Asset inventory
    - Conduit analysis
    - IEC 62443-3-3 requirements
    - Recommended security controls
    - Validation results
    - Policy violations
    """
    import base64
    import io
    from datetime import datetime as dt

    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle
        from reportlab.platypus import Table as RLTable
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="PDF export requires the 'reportlab' package. "
            "Install with: pip install reportlab",
        )

    from induform.engine.gap_analysis import analyze_gaps
    from induform.engine.resolver import resolve_security_controls
    from induform.engine.validator import validate_project
    from induform.iec62443.requirements import get_requirements_for_level

    project_repo = ProjectRepository(db)

    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project = await project_repo.to_pydantic(project_db)

    # Calculate stats
    total_zones = len(project.zones)
    total_assets = sum(len(z.assets) for z in project.zones)
    total_conduits = len(project.conduits)

    # Get risk assessment
    vuln_data = await _load_vulnerability_data(db, project_db.id)
    risk_result = assess_risk(project, vulnerability_data=vuln_data)

    # Get gap analysis
    gap_report = analyze_gaps(project)

    # Get policy violations
    violations = evaluate_policies(project)
    critical_violations = [v for v in violations if v.severity == PolicySeverity.CRITICAL]
    high_violations = [v for v in violations if v.severity == PolicySeverity.HIGH]

    # Get security controls
    security_controls = resolve_security_controls(project)

    # Get validation results
    validation_report = validate_project(project)

    # Max SL-T across all zones
    max_sl = max((z.security_level_target for z in project.zones), default=1)
    applicable_requirements = get_requirements_for_level(max_sl)

    # Page number callback
    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#64748b"))
        page_num = canvas.getPageNumber()
        canvas.drawRightString(letter[0] - 0.75 * inch, 0.5 * inch, f"Page {page_num}")
        canvas.drawString(0.75 * inch, 0.5 * inch, f"InduForm - {project_db.name}")
        canvas.restoreState()

    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter, topMargin=0.75 * inch, bottomMargin=0.75 * inch
    )
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Heading1"], fontSize=24, spaceAfter=20, alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=14,
        spaceAfter=10,
        textColor=colors.HexColor("#1e40af"),
    )
    subheading_style = ParagraphStyle(
        "CustomSubheading",
        parent=styles["Heading3"],
        fontSize=11,
        spaceAfter=6,
        textColor=colors.HexColor("#334155"),
    )
    normal_style = styles["Normal"]

    # Standard table style
    def make_table_style(header_color="#1e40af", alt_row_color="#f8fafc"):
        return TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_color)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#94a3b8")),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor(alt_row_color)],
                ),
                ("TOPPADDING", (0, 1), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
            ]
        )

    # --- Title Page ---
    story.append(Spacer(1, 2 * inch))
    story.append(Paragraph("Security Assessment Report", title_style))
    story.append(
        Paragraph(
            f"<b>{project_db.name}</b>",
            ParagraphStyle(
                "ProjectName", parent=styles["Heading2"], fontSize=18, alignment=TA_CENTER
            ),
        )
    )
    story.append(Spacer(1, 0.5 * inch))
    story.append(
        Paragraph(
            f"Generated: {dt.now().strftime('%Y-%m-%d %H:%M')}",
            ParagraphStyle("Date", parent=normal_style, alignment=TA_CENTER),
        )
    )
    story.append(
        Paragraph(
            "Standard: IEC 62443",
            ParagraphStyle("Standard", parent=normal_style, alignment=TA_CENTER),
        )
    )
    story.append(
        Paragraph(
            f"Maximum Security Level Target: SL-{max_sl}",
            ParagraphStyle("SLInfo", parent=normal_style, alignment=TA_CENTER),
        )
    )
    story.append(PageBreak())

    # --- Table of Contents ---
    story.append(Paragraph("Table of Contents", heading_style))
    story.append(Spacer(1, 0.2 * inch))
    toc_items = [
        "1. Executive Summary",
        "2. Zone Inventory",
        "3. Asset Inventory",
        "4. Conduit Summary",
        "5. Risk Assessment",
        "6. Gap Analysis",
        "7. Vulnerability Summary",
        "8. IEC 62443-3-3 Applicable Requirements",
        "9. Recommended Security Controls",
        "10. Validation Results",
        "11. Policy Violations",
        "12. Network Topology",
        "13. Attack Path Analysis",
    ]
    for item in toc_items:
        story.append(
            Paragraph(
                item,
                ParagraphStyle(
                    "TOCItem", parent=normal_style, fontSize=11, spaceAfter=6, leftIndent=20
                ),
            )
        )
    story.append(PageBreak())

    # --- 1. Executive Summary ---
    story.append(Paragraph("1. Executive Summary", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    summary_data = [
        ["Metric", "Value"],
        ["Total Zones", str(total_zones)],
        ["Total Assets", str(total_assets)],
        ["Total Conduits", str(total_conduits)],
        ["Maximum SL-T", f"SL-{max_sl}"],
        ["Overall Risk Score", f"{risk_result.overall_score:.0f}/100"],
        ["Risk Level", risk_result.overall_level.value.upper()],
        ["Critical Violations", str(len(critical_violations))],
        ["High Violations", str(len(high_violations))],
        ["Validation Errors", str(validation_report.error_count)],
        ["Validation Warnings", str(validation_report.warning_count)],
        ["Compliance Score", f"{gap_report.overall_compliance:.1f}%"],
    ]

    t = RLTable(summary_data, colWidths=[3 * inch, 2 * inch])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f1f5f9")),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#cbd5e1")),
                ("FONTSIZE", (0, 1), (-1, -1), 10),
                ("TOPPADDING", (0, 1), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))

    if project_db.description:
        story.append(Paragraph(f"<b>Description:</b> {project_db.description}", normal_style))
        story.append(Spacer(1, 0.2 * inch))

    # --- 2. Zone Inventory ---
    story.append(Paragraph("2. Zone Inventory", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    if project.zones:
        zone_data = [["Zone ID", "Name", "Type", "SL-T", "Assets"]]
        for zone in project.zones:
            zone_type = zone.type.value if hasattr(zone.type, "value") else str(zone.type)
            zone_data.append(
                [
                    zone.id,
                    zone.name[:25] + "..." if len(zone.name) > 25 else zone.name,
                    zone_type,
                    str(zone.security_level_target),
                    str(len(zone.assets)),
                ]
            )

        zt = RLTable(
            zone_data, colWidths=[1.2 * inch, 2 * inch, 1.2 * inch, 0.6 * inch, 0.8 * inch]
        )
        zt.setStyle(make_table_style())
        story.append(zt)
    else:
        story.append(Paragraph("No zones defined.", normal_style))

    story.append(Spacer(1, 0.3 * inch))

    # --- 3. Asset Inventory ---
    story.append(Paragraph("3. Asset Inventory", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    all_assets = []
    for zone in project.zones:
        for asset in zone.assets:
            asset_type = asset.type.value if hasattr(asset.type, "value") else str(asset.type)
            all_assets.append(
                (zone.id, asset.name, asset_type, asset.ip_address or "-", asset.criticality)
            )

    if all_assets:
        asset_data = [["Zone", "Asset Name", "Type", "IP Address", "Criticality"]]
        for zone_id, name, atype, ip, crit in all_assets[:30]:
            asset_data.append(
                [
                    zone_id,
                    name[:20] + "..." if len(name) > 20 else name,
                    atype,
                    ip,
                    str(crit),
                ]
            )

        at = RLTable(
            asset_data, colWidths=[1.1 * inch, 1.6 * inch, 1 * inch, 1.2 * inch, 0.8 * inch]
        )
        at.setStyle(make_table_style())
        story.append(at)
        if len(all_assets) > 30:
            story.append(Paragraph(f"<i>Showing 30 of {len(all_assets)} assets</i>", normal_style))
    else:
        story.append(Paragraph("No assets defined.", normal_style))

    story.append(Spacer(1, 0.3 * inch))

    # --- 4. Conduit Summary ---
    story.append(Paragraph("4. Conduit Summary", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    if project.conduits:
        conduit_data = [["ID", "From", "To", "SL-R", "Inspection"]]
        for conduit in project.conduits[:20]:
            conduit_data.append(
                [
                    conduit.id[:15] + "..." if len(conduit.id) > 15 else conduit.id,
                    conduit.from_zone,
                    conduit.to_zone,
                    str(conduit.security_level_required or "-"),
                    "Yes" if conduit.requires_inspection else "No",
                ]
            )

        ct = RLTable(
            conduit_data, colWidths=[1.3 * inch, 1.3 * inch, 1.3 * inch, 0.6 * inch, 0.9 * inch]
        )
        ct.setStyle(make_table_style())
        story.append(ct)
        if len(project.conduits) > 20:
            story.append(
                Paragraph(f"<i>Showing 20 of {len(project.conduits)} conduits</i>", normal_style)
            )
    else:
        story.append(Paragraph("No conduits defined.", normal_style))

    story.append(Spacer(1, 0.3 * inch))

    # --- 5. Risk Assessment ---
    story.append(PageBreak())
    story.append(Paragraph("5. Risk Assessment", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    story.append(
        Paragraph(
            f"Overall risk score: <b>{risk_result.overall_score:.0f}/100</b> "
            f"(Level: <b>{risk_result.overall_level.value.upper()}</b>). "
            f"Assessment covers {total_zones} zone(s) with {total_assets} asset(s) "
            f"across {total_conduits} conduit connection(s).",
            normal_style,
        )
    )
    story.append(Spacer(1, 0.15 * inch))

    if risk_result.zone_risks:
        story.append(Paragraph("Per-Zone Risk Breakdown", subheading_style))
        risk_table_data = [
            ["Zone", "Score", "Level", "SL Base", "Asset Crit", "Exposure", "SL Gap", "Vuln"]
        ]
        for zone_id, zr in risk_result.zone_risks.items():
            zone_label = zone_id[:15] + "..." if len(zone_id) > 15 else zone_id
            risk_table_data.append(
                [
                    zone_label,
                    f"{zr.score:.0f}",
                    zr.level.value.upper(),
                    f"{zr.factors.sl_base_risk:.0f}",
                    f"{zr.factors.asset_criticality_risk:.0f}",
                    f"{zr.factors.exposure_risk:.0f}",
                    f"{zr.factors.sl_gap_risk:.0f}",
                    f"{zr.factors.vulnerability_risk:.0f}",
                ]
            )

        rsk_t = RLTable(
            risk_table_data,
            colWidths=[
                1.1 * inch,
                0.5 * inch,
                0.7 * inch,
                0.6 * inch,
                0.7 * inch,
                0.7 * inch,
                0.6 * inch,
                0.5 * inch,
            ],
        )
        rsk_t.setStyle(make_table_style("#dc2626"))
        story.append(rsk_t)

    story.append(Spacer(1, 0.15 * inch))

    if risk_result.recommendations:
        story.append(Paragraph("Risk Recommendations", subheading_style))
        for rec in risk_result.recommendations[:10]:
            bullet_text = f"\u2022 {rec}"
            story.append(
                Paragraph(
                    bullet_text,
                    ParagraphStyle(
                        "RiskBullet", parent=normal_style, fontSize=9, leftIndent=15, spaceAfter=3
                    ),
                )
            )

    story.append(Spacer(1, 0.3 * inch))

    # Risk Matrix visualization
    story.append(Paragraph("Risk Matrix", subheading_style))
    risk_matrix_drawing = _draw_risk_matrix(risk_result, project.zones)
    try:
        from reportlab.graphics import renderPDF
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="PDF export requires the 'reportlab' package. "
            "Install with: pip install reportlab",
        )

    story.append(renderPDF.GraphicsFlowable(risk_matrix_drawing))
    story.append(Spacer(1, 0.3 * inch))

    # --- 6. Gap Analysis ---
    story.append(PageBreak())
    story.append(Paragraph("6. Gap Analysis", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    story.append(
        Paragraph(
            f"Overall IEC 62443-3-3 compliance: <b>{gap_report.overall_compliance:.1f}%</b>. "
            f"Controls met: <b>{gap_report.summary.get('met', 0)}</b>, "
            f"partial: <b>{gap_report.summary.get('partial', 0)}</b>, "
            f"unmet: <b>{gap_report.summary.get('unmet', 0)}</b>.",
            normal_style,
        )
    )
    story.append(Spacer(1, 0.15 * inch))

    if gap_report.zones:
        story.append(Paragraph("Per-Zone Compliance", subheading_style))
        gap_table_data = [
            ["Zone", "Type", "SL-T", "Controls", "Met", "Partial", "Unmet", "Compliance%"]
        ]
        for za in gap_report.zones:
            gap_table_data.append(
                [
                    za.zone_name[:15] + "..." if len(za.zone_name) > 15 else za.zone_name,
                    za.zone_type,
                    str(za.security_level_target),
                    str(za.total_controls),
                    str(za.met_controls),
                    str(za.partial_controls),
                    str(za.unmet_controls),
                    f"{za.compliance_percentage:.0f}%",
                ]
            )

        gap_t = RLTable(
            gap_table_data,
            colWidths=[
                1.1 * inch,
                0.8 * inch,
                0.4 * inch,
                0.6 * inch,
                0.5 * inch,
                0.5 * inch,
                0.5 * inch,
                0.8 * inch,
            ],
        )
        gap_t.setStyle(make_table_style("#0f766e"))
        story.append(gap_t)

    story.append(Spacer(1, 0.15 * inch))

    if gap_report.priority_remediations:
        story.append(Paragraph("Priority Remediations", subheading_style))
        for rem in gap_report.priority_remediations[:8]:
            bullet_text = f"\u2022 {rem}"
            story.append(
                Paragraph(
                    bullet_text,
                    ParagraphStyle(
                        "GapBullet", parent=normal_style, fontSize=9, leftIndent=15, spaceAfter=3
                    ),
                )
            )

    story.append(Spacer(1, 0.3 * inch))

    # --- 7. Vulnerability Summary ---
    story.append(PageBreak())
    story.append(Paragraph("7. Vulnerability Summary", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    # Flatten vuln_data into a list with zone info
    all_vulns: list[tuple[str, VulnInfo]] = []
    for zone_id, zone_vulns in vuln_data.items():
        for v in zone_vulns:
            all_vulns.append((zone_id, v))

    sev_counts: dict[str, int] = {}
    for _zid, v in all_vulns:
        sev_counts[v.severity] = sev_counts.get(v.severity, 0) + 1

    sev_breakdown = (
        ", ".join(f"{sev}: {count}" for sev, count in sorted(sev_counts.items())) or "none"
    )

    story.append(
        Paragraph(
            f"Total vulnerabilities: <b>{len(all_vulns)}</b>. "
            f"Breakdown by severity: {sev_breakdown}.",
            normal_style,
        )
    )
    story.append(Spacer(1, 0.15 * inch))

    if all_vulns:
        story.append(Paragraph("Top Vulnerabilities", subheading_style))
        vuln_table_data = [["CVE ID", "Zone", "Severity", "CVSS", "Status"]]
        # Sort by CVSS descending (None treated as 0)
        sorted_vulns = sorted(all_vulns, key=lambda x: x[1].cvss_score or 0, reverse=True)
        for zone_id, v in sorted_vulns[:20]:
            vuln_table_data.append(
                [
                    v.cve_id,
                    zone_id[:15] + "..." if len(zone_id) > 15 else zone_id,
                    v.severity.upper(),
                    f"{v.cvss_score:.1f}" if v.cvss_score is not None else "-",
                    v.status,
                ]
            )

        vul_t = RLTable(
            vuln_table_data,
            colWidths=[1.3 * inch, 1.3 * inch, 0.8 * inch, 0.6 * inch, 0.8 * inch],
        )
        vul_t.setStyle(make_table_style("#b91c1c", "#fef2f2"))
        story.append(vul_t)
        if len(all_vulns) > 20:
            story.append(
                Paragraph(f"<i>Showing 20 of {len(all_vulns)} vulnerabilities</i>", normal_style)
            )
    else:
        story.append(Paragraph("No vulnerabilities recorded.", normal_style))

    story.append(Spacer(1, 0.3 * inch))

    # --- 8. IEC 62443-3-3 Applicable Requirements ---
    story.append(PageBreak())
    story.append(Paragraph("8. IEC 62443-3-3 Applicable Requirements", heading_style))
    story.append(Spacer(1, 0.1 * inch))
    story.append(
        Paragraph(
            f"Based on the maximum Security Level Target (SL-{max_sl}) in this project, "
            f"the following {len(applicable_requirements)} requirements from IEC 62443-3-3 apply:",
            normal_style,
        )
    )
    story.append(Spacer(1, 0.2 * inch))

    if applicable_requirements:
        req_data = [["SR ID", "Name", "FR Category", "Min SL"]]
        for req in applicable_requirements:
            fr_short = (
                req.foundational_requirement.split(" - ")[0]
                if " - " in req.foundational_requirement
                else req.foundational_requirement
            )
            req_data.append(
                [
                    req.id,
                    req.name[:30] + "..." if len(req.name) > 30 else req.name,
                    fr_short,
                    f"SL-{req.minimum_sl}",
                ]
            )

        rt = RLTable(req_data, colWidths=[0.7 * inch, 2.5 * inch, 1.5 * inch, 0.6 * inch])
        rt.setStyle(make_table_style("#0f766e"))
        story.append(rt)

    story.append(Spacer(1, 0.3 * inch))

    # --- 9. Recommended Security Controls ---
    story.append(Paragraph("9. Recommended Security Controls", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    # Global controls
    global_controls = security_controls.get("global_controls", [])
    if global_controls:
        story.append(Paragraph("Project-Wide Controls", subheading_style))
        gc_data = [["Control", "Description", "Priority"]]
        for ctrl in global_controls:
            gc_data.append(
                [
                    ctrl["control"],
                    ctrl["description"][:45] + "..."
                    if len(ctrl["description"]) > 45
                    else ctrl["description"],
                    str(ctrl["priority"]),
                ]
            )

        gt = RLTable(gc_data, colWidths=[1.5 * inch, 3.5 * inch, 0.7 * inch])
        gt.setStyle(make_table_style("#7c3aed"))
        story.append(gt)
        story.append(Spacer(1, 0.2 * inch))

    # Zone-level control summary
    zone_profiles = security_controls.get("zone_profiles", [])
    if zone_profiles:
        story.append(Paragraph("Zone Security Profiles", subheading_style))
        zp_data = [["Zone", "SL-T", "Requirements", "Top Controls"]]
        for zp in zone_profiles[:15]:
            top_controls = [c["requirement_id"] for c in zp.get("recommended_controls", [])[:3]]
            zp_data.append(
                [
                    zp["zone_id"],
                    f"SL-{zp['security_level_target']}",
                    str(len(zp.get("applicable_requirements", []))),
                    ", ".join(top_controls) if top_controls else "-",
                ]
            )

        zpt = RLTable(zp_data, colWidths=[1.2 * inch, 0.6 * inch, 1 * inch, 2.8 * inch])
        zpt.setStyle(make_table_style("#7c3aed"))
        story.append(zpt)

    story.append(Spacer(1, 0.3 * inch))

    # --- 10. Validation Results ---
    story.append(PageBreak())
    story.append(Paragraph("10. Validation Results", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    if validation_report.results:
        val_summary = (
            f"Validation found <b>{validation_report.error_count}</b> error(s), "
            f"<b>{validation_report.warning_count}</b> warning(s), and "
            f"<b>{validation_report.info_count}</b> informational finding(s)."
        )
        story.append(Paragraph(val_summary, normal_style))
        story.append(Spacer(1, 0.15 * inch))

        val_data = [["Severity", "Code", "Message"]]
        for r in validation_report.results[:20]:
            val_data.append(
                [
                    r.severity.value.upper(),
                    r.code,
                    r.message[:55] + "..." if len(r.message) > 55 else r.message,
                ]
            )

        vrt = RLTable(val_data, colWidths=[0.8 * inch, 1.8 * inch, 3.1 * inch])
        err_style = make_table_style("#b91c1c", "#fef2f2")
        vrt.setStyle(err_style)
        story.append(vrt)
        if len(validation_report.results) > 20:
            story.append(
                Paragraph(
                    f"<i>Showing 20 of {len(validation_report.results)} findings</i>", normal_style
                )
            )
    else:
        story.append(
            Paragraph(
                "<b>No validation issues found.</b> Configuration passes all checks.", normal_style
            )
        )

    story.append(Spacer(1, 0.3 * inch))

    # --- 11. Policy Violations ---
    story.append(Paragraph("11. Policy Violations", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    if violations:
        viol_data = [["Severity", "Rule", "Affected", "Message"]]
        for v in violations[:15]:
            affected = ", ".join(v.affected_entities[:2])
            if len(v.affected_entities) > 2:
                affected += "..."
            viol_data.append(
                [
                    v.severity.value.upper(),
                    v.rule_id,
                    affected or "-",
                    v.message[:40] + "..." if len(v.message) > 40 else v.message,
                ]
            )

        vt = RLTable(viol_data, colWidths=[0.7 * inch, 0.8 * inch, 1.3 * inch, 2.9 * inch])
        vt.setStyle(make_table_style("#dc2626", "#fef2f2"))
        story.append(vt)
        if len(violations) > 15:
            story.append(
                Paragraph(f"<i>Showing 15 of {len(violations)} violations</i>", normal_style)
            )
    else:
        story.append(
            Paragraph("<b>No policy violations found.</b> All policy rules pass.", normal_style)
        )

    # --- 12. Network Topology ---
    story.append(PageBreak())
    story.append(Paragraph("12. Network Topology", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    if project.zones:
        story.append(
            Paragraph(
                f"Zone/conduit topology diagram showing {total_zones} zone(s) and "
                f"{total_conduits} conduit connection(s), arranged by Purdue model layers.",
                normal_style,
            )
        )
        story.append(Spacer(1, 0.15 * inch))
        topo_drawing = _draw_topology_diagram(project.zones, project.conduits)
        story.append(renderPDF.GraphicsFlowable(topo_drawing))
    else:
        story.append(Paragraph("No zones defined — topology diagram unavailable.", normal_style))

    story.append(Spacer(1, 0.3 * inch))

    # --- 13. Attack Path Analysis ---
    story.append(PageBreak())
    story.append(Paragraph("13. Attack Path Analysis", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    from induform.engine.attack_path import analyze_attack_paths

    attack_analysis = analyze_attack_paths(project)

    story.append(Paragraph(attack_analysis.summary, normal_style))
    story.append(Spacer(1, 0.15 * inch))

    if attack_analysis.entry_points:
        entry_text = ", ".join(attack_analysis.entry_points)
        story.append(Paragraph(f"<b>Entry points:</b> {entry_text}", normal_style))

    if attack_analysis.high_value_targets:
        target_text = ", ".join(attack_analysis.high_value_targets)
        story.append(Paragraph(f"<b>High-value targets:</b> {target_text}", normal_style))

    story.append(Spacer(1, 0.15 * inch))

    if attack_analysis.paths:
        # Path summary table
        story.append(Paragraph("Attack Path Summary", subheading_style))
        path_data = [["Entry", "Target", "Steps", "Risk Score", "Risk Level"]]
        for path in attack_analysis.paths:
            entry_name = path.entry_zone_name[:15]
            target_name = path.target_zone_name[:15]
            path_data.append(
                [
                    entry_name,
                    target_name,
                    str(len(path.steps)),
                    f"{path.risk_score:.0f}/100",
                    path.risk_level.upper(),
                ]
            )

        apt = RLTable(
            path_data,
            colWidths=[1.2 * inch, 1.2 * inch, 0.6 * inch, 0.9 * inch, 0.9 * inch],
        )
        apt.setStyle(make_table_style("#b91c1c", "#fef2f2"))
        story.append(apt)
        story.append(Spacer(1, 0.2 * inch))

        # Detailed breakdown for top 5 paths
        story.append(Paragraph("Detailed Path Analysis (Top 5)", subheading_style))
        for idx, path in enumerate(attack_analysis.paths[:5], 1):
            story.append(
                Paragraph(
                    f"<b>Path {idx}:</b> {path.entry_zone_name} → {path.target_zone_name} "
                    f"(Risk: {path.risk_score:.0f}, {path.risk_level.upper()})",
                    ParagraphStyle(
                        "PathTitle",
                        parent=normal_style,
                        fontSize=10,
                        spaceAfter=4,
                        textColor=colors.HexColor("#1e40af"),
                    ),
                )
            )
            story.append(
                Paragraph(
                    f"<i>Reason:</i> {path.target_reason}",
                    ParagraphStyle(
                        "PathReason",
                        parent=normal_style,
                        fontSize=9,
                        leftIndent=15,
                        spaceAfter=3,
                    ),
                )
            )

            # Step chain
            step_parts = [path.entry_zone_name]
            for step in path.steps:
                step_parts.append(f"→ {step.to_zone_name}")
            chain_text = " ".join(step_parts)
            story.append(
                Paragraph(
                    f"<b>Route:</b> {chain_text}",
                    ParagraphStyle(
                        "PathChain",
                        parent=normal_style,
                        fontSize=9,
                        leftIndent=15,
                        spaceAfter=3,
                    ),
                )
            )

            # Weaknesses
            all_weaknesses = []
            for step in path.steps:
                for w in step.weaknesses:
                    all_weaknesses.append(w)

            if all_weaknesses:
                for w in all_weaknesses[:5]:
                    story.append(
                        Paragraph(
                            f"\u2022 {w.description}",
                            ParagraphStyle(
                                "WeakBullet",
                                parent=normal_style,
                                fontSize=8,
                                leftIndent=25,
                                spaceAfter=2,
                            ),
                        )
                    )
                    story.append(
                        Paragraph(
                            f"  ↳ {w.remediation}",
                            ParagraphStyle(
                                "WeakRemediation",
                                parent=normal_style,
                                fontSize=8,
                                leftIndent=35,
                                spaceAfter=2,
                                textColor=colors.HexColor("#059669"),
                            ),
                        )
                    )

            story.append(Spacer(1, 0.15 * inch))
    else:
        story.append(
            Paragraph(
                "No attack paths identified between entry points and high-value targets.",
                normal_style,
            )
        )

    # Build PDF with page numbers
    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)

    # Return as base64 encoded string
    pdf_data = base64.b64encode(buffer.getvalue()).decode("utf-8")
    filename = f"{project_db.name.lower().replace(' ', '_')}_report.pdf"

    return {"pdf_base64": pdf_data, "filename": filename}


# Gap Analysis endpoint


@router.get("/{project_id}/gap-analysis", response_model=GapAnalysisReport)
async def get_gap_analysis(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> GapAnalysisReport:
    """Perform IEC 62443-3-3 compliance gap analysis for a project.

    Maps foundational requirements (FR1-FR7) to system requirements (SRs)
    and assesses which controls are met, partially met, or unmet for each zone
    based on zone configuration, assets, and conduit settings.
    """
    project_repo = ProjectRepository(db)

    # Check permission (viewer can access gap analysis)
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Convert to Pydantic model and run analysis
    project = await project_repo.to_pydantic(project_db)
    report = analyze_gaps(project)

    return report


# Attack Path Analysis endpoint


@router.post("/{project_id}/attack-paths", response_model=AttackPathAnalysis)
async def project_attack_paths(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> AttackPathAnalysis:
    """Analyze attack paths for a project."""
    project_repo = ProjectRepository(db)
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )
    project_db = await project_repo.get_by_id(project_id)
    if not project_db:
        raise HTTPException(status_code=404, detail="Project not found")
    project = await project_repo.to_pydantic(project_db)
    return analyze_attack_paths(project)


# Project comparison endpoint
class CompareProjectsRequest(BaseModel):
    """Request to compare two projects."""

    project_a_id: str
    project_b_id: str


@router.post("/compare", response_model=ComparisonResult)
async def compare_projects(
    request: CompareProjectsRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ComparisonResult:
    """
    Compare two projects and return the differences.

    Returns:
    - zones: added, removed, modified zones
    - assets: added, removed, modified assets
    - conduits: added, removed, modified conduits
    - summary: count of changes by category
    """
    project_repo = ProjectRepository(db)

    # Check access to both projects
    has_access_a = await check_project_permission(
        db, request.project_a_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    has_access_b = await check_project_permission(
        db, request.project_b_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )

    if not has_access_a or not has_access_b:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="One or both projects not found",
        )

    project_a_db = await project_repo.get_by_id(request.project_a_id)
    project_b_db = await project_repo.get_by_id(request.project_b_id)

    if not project_a_db or not project_b_db:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="One or both projects not found",
        )

    project_a = await project_repo.to_pydantic(project_a_db)
    project_b = await project_repo.to_pydantic(project_b_db)

    # Compare zones
    zones_a = {z.id: z for z in project_a.zones}
    zones_b = {z.id: z for z in project_b.zones}

    added_zones = []
    removed_zones = []
    modified_zones = []

    for zone_id in set(zones_a.keys()) | set(zones_b.keys()):
        if zone_id not in zones_a:
            z = zones_b[zone_id]
            added_zones.append(
                {
                    "id": zone_id,
                    "name": z.name,
                    "type": z.type,
                    "security_level_target": z.security_level_target,
                }
            )
        elif zone_id not in zones_b:
            z = zones_a[zone_id]
            removed_zones.append(
                {
                    "id": zone_id,
                    "name": z.name,
                    "type": z.type,
                    "security_level_target": z.security_level_target,
                }
            )
        else:
            za = zones_a[zone_id]
            zb = zones_b[zone_id]
            changes = {}
            if za.name != zb.name:
                changes["name"] = {"from": za.name, "to": zb.name}
            if za.type != zb.type:
                changes["type"] = {"from": za.type, "to": zb.type}
            if za.security_level_target != zb.security_level_target:
                changes["security_level_target"] = {
                    "from": za.security_level_target,
                    "to": zb.security_level_target,
                }
            if len(za.assets) != len(zb.assets):
                changes["asset_count"] = {"from": len(za.assets), "to": len(zb.assets)}

            if changes:
                modified_zones.append(
                    {
                        "id": zone_id,
                        "name": zb.name,
                        "changes": changes,
                    }
                )

    # Compare conduits
    conduits_a = {c.id: c for c in project_a.conduits}
    conduits_b = {c.id: c for c in project_b.conduits}

    added_conduits = []
    removed_conduits = []
    modified_conduits = []

    for conduit_id in set(conduits_a.keys()) | set(conduits_b.keys()):
        if conduit_id not in conduits_a:
            c = conduits_b[conduit_id]
            added_conduits.append(
                {
                    "id": conduit_id,
                    "from_zone": c.from_zone,
                    "to_zone": c.to_zone,
                }
            )
        elif conduit_id not in conduits_b:
            c = conduits_a[conduit_id]
            removed_conduits.append(
                {
                    "id": conduit_id,
                    "from_zone": c.from_zone,
                    "to_zone": c.to_zone,
                }
            )
        else:
            ca = conduits_a[conduit_id]
            cb = conduits_b[conduit_id]
            changes = {}
            if ca.from_zone != cb.from_zone:
                changes["from_zone"] = {"from": ca.from_zone, "to": cb.from_zone}
            if ca.to_zone != cb.to_zone:
                changes["to_zone"] = {"from": ca.to_zone, "to": cb.to_zone}
            if ca.security_level_required != cb.security_level_required:
                changes["security_level_required"] = {
                    "from": ca.security_level_required,
                    "to": cb.security_level_required,
                }
            if len(ca.flows) != len(cb.flows):
                changes["flow_count"] = {"from": len(ca.flows), "to": len(cb.flows)}

            if changes:
                modified_conduits.append(
                    {
                        "id": conduit_id,
                        "changes": changes,
                    }
                )

    # Compare assets across all zones
    all_assets_a = {}
    all_assets_b = {}
    for zone in project_a.zones:
        for asset in zone.assets:
            all_assets_a[f"{zone.id}:{asset.id}"] = (zone.id, asset)
    for zone in project_b.zones:
        for asset in zone.assets:
            all_assets_b[f"{zone.id}:{asset.id}"] = (zone.id, asset)

    added_assets = []
    removed_assets = []
    modified_assets = []

    for asset_key in set(all_assets_a.keys()) | set(all_assets_b.keys()):
        if asset_key not in all_assets_a:
            zone_id, asset = all_assets_b[asset_key]
            added_assets.append(
                {
                    "zone_id": zone_id,
                    "id": asset.id,
                    "name": asset.name,
                    "type": asset.type,
                }
            )
        elif asset_key not in all_assets_b:
            zone_id, asset = all_assets_a[asset_key]
            removed_assets.append(
                {
                    "zone_id": zone_id,
                    "id": asset.id,
                    "name": asset.name,
                    "type": asset.type,
                }
            )
        else:
            zone_id_a, aa = all_assets_a[asset_key]
            zone_id_b, ab = all_assets_b[asset_key]
            changes = {}
            if aa.name != ab.name:
                changes["name"] = {"from": aa.name, "to": ab.name}
            if aa.type != ab.type:
                changes["type"] = {"from": aa.type, "to": ab.type}
            if aa.ip_address != ab.ip_address:
                changes["ip_address"] = {"from": aa.ip_address, "to": ab.ip_address}
            if aa.criticality != ab.criticality:
                changes["criticality"] = {"from": aa.criticality, "to": ab.criticality}

            if changes:
                modified_assets.append(
                    {
                        "zone_id": zone_id_b,
                        "id": ab.id,
                        "name": ab.name,
                        "changes": changes,
                    }
                )

    return ComparisonResult(
        zones={
            "added": added_zones,
            "removed": removed_zones,
            "modified": modified_zones,
        },
        assets={
            "added": added_assets,
            "removed": removed_assets,
            "modified": modified_assets,
        },
        conduits={
            "added": added_conduits,
            "removed": removed_conduits,
            "modified": modified_conduits,
        },
        summary={
            "zones_added": len(added_zones),
            "zones_removed": len(removed_zones),
            "zones_modified": len(modified_zones),
            "assets_added": len(added_assets),
            "assets_removed": len(removed_assets),
            "assets_modified": len(modified_assets),
            "conduits_added": len(added_conduits),
            "conduits_removed": len(removed_conduits),
            "conduits_modified": len(modified_conduits),
        },
    )


# --- Metrics / Analytics ---

_METRICS_THROTTLE_SECONDS = 300  # 5 minutes


async def _record_metrics_snapshot(
    db: AsyncSession,
    project_id: str,
    project: Project,
) -> None:
    """Record a time-series metrics snapshot for a project.

    Throttled to max 1 snapshot per 5 minutes per project to avoid flooding.
    """
    from datetime import timedelta

    from sqlalchemy import select as sa_select

    from induform.db.models import MetricsSnapshot
    from induform.engine.validator import validate_project

    # Check throttle: skip if a snapshot was recorded recently
    cutoff = datetime.utcnow() - timedelta(seconds=_METRICS_THROTTLE_SECONDS)
    recent_query = (
        sa_select(MetricsSnapshot.id)
        .where(MetricsSnapshot.project_id == project_id)
        .where(MetricsSnapshot.recorded_at >= cutoff)
        .limit(1)
    )
    result = await db.execute(recent_query)
    if result.scalar_one_or_none() is not None:
        return  # Throttled

    # Calculate metrics
    zone_count = len(project.zones)
    asset_count = sum(len(z.assets) for z in project.zones)
    conduit_count = len(project.conduits)

    # Compliance score from policy violations
    compliance_score = 100.0
    try:
        enabled_standards = project.project.compliance_standards or None
        violations = evaluate_policies(project, enabled_standards=enabled_standards)
        if violations:
            deduction = 0.0
            for v in violations:
                if v.severity == PolicySeverity.CRITICAL:
                    deduction += 25
                elif v.severity == PolicySeverity.HIGH:
                    deduction += 15
                elif v.severity == PolicySeverity.MEDIUM:
                    deduction += 8
                else:
                    deduction += 3
            compliance_score = max(0.0, 100.0 - deduction)
    except Exception:
        pass

    # Risk score
    risk_score = 0.0
    try:
        vuln_data = await _load_vulnerability_data(db, project_id)
        risk_assessment = assess_risk(project, vulnerability_data=vuln_data)
        risk_score = risk_assessment.overall_score
    except Exception:
        pass

    # Validation results
    error_count = 0
    warning_count = 0
    try:
        validation_report = validate_project(project)
        error_count = validation_report.error_count
        warning_count = validation_report.warning_count
    except Exception:
        pass

    snapshot = MetricsSnapshot(
        project_id=project_id,
        zone_count=zone_count,
        asset_count=asset_count,
        conduit_count=conduit_count,
        compliance_score=compliance_score,
        risk_score=risk_score,
        error_count=error_count,
        warning_count=warning_count,
    )
    db.add(snapshot)
    await db.flush()


# Analytics response schemas


class MetricsDataPoint(BaseModel):
    """Single metrics data point."""

    recorded_at: datetime
    zone_count: int
    asset_count: int
    conduit_count: int
    compliance_score: float
    risk_score: float
    error_count: int
    warning_count: int


class TrendDirection(BaseModel):
    """Trend direction for a metric."""

    value: float
    direction: str  # "up", "down", "stable"
    change: float  # absolute change


class AnalyticsSummary(BaseModel):
    """Summary analytics for a project."""

    current: MetricsDataPoint | None
    compliance_trend: TrendDirection | None
    risk_trend: TrendDirection | None
    zone_count_trend: TrendDirection | None
    asset_count_trend: TrendDirection | None
    min_compliance: float | None
    max_compliance: float | None
    min_risk: float | None
    max_risk: float | None
    snapshot_count: int


@router.get("/{project_id}/analytics", response_model=list[MetricsDataPoint])
async def get_project_analytics(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    days: int = Query(30, ge=1, le=365, description="Number of days of history"),
) -> list[MetricsDataPoint]:
    """Return time-series analytics data for a project."""
    from datetime import timedelta

    from sqlalchemy import select as sa_select

    from induform.db.models import MetricsSnapshot

    # Check permission
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    cutoff = datetime.utcnow() - timedelta(days=days)
    query = (
        sa_select(MetricsSnapshot)
        .where(MetricsSnapshot.project_id == project_id)
        .where(MetricsSnapshot.recorded_at >= cutoff)
        .order_by(MetricsSnapshot.recorded_at.asc())
    )
    result = await db.execute(query)
    snapshots = result.scalars().all()

    return [
        MetricsDataPoint(
            recorded_at=s.recorded_at,
            zone_count=s.zone_count,
            asset_count=s.asset_count,
            conduit_count=s.conduit_count,
            compliance_score=s.compliance_score,
            risk_score=s.risk_score,
            error_count=s.error_count,
            warning_count=s.warning_count,
        )
        for s in snapshots
    ]


@router.get("/{project_id}/analytics/summary", response_model=AnalyticsSummary)
async def get_project_analytics_summary(
    project_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    days: int = Query(30, ge=1, le=365, description="Number of days of history"),
) -> AnalyticsSummary:
    """Return summary analytics with trends for a project."""
    from datetime import timedelta

    from sqlalchemy import select as sa_select

    from induform.db.models import MetricsSnapshot

    # Check permission
    has_access = await check_project_permission(
        db, project_id, current_user.id, Permission.VIEWER, is_admin=current_user.is_admin
    )
    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    cutoff = datetime.utcnow() - timedelta(days=days)
    query = (
        sa_select(MetricsSnapshot)
        .where(MetricsSnapshot.project_id == project_id)
        .where(MetricsSnapshot.recorded_at >= cutoff)
        .order_by(MetricsSnapshot.recorded_at.asc())
    )
    result = await db.execute(query)
    snapshots = list(result.scalars().all())

    if not snapshots:
        return AnalyticsSummary(
            current=None,
            compliance_trend=None,
            risk_trend=None,
            zone_count_trend=None,
            asset_count_trend=None,
            min_compliance=None,
            max_compliance=None,
            min_risk=None,
            max_risk=None,
            snapshot_count=0,
        )

    current = snapshots[-1]
    current_point = MetricsDataPoint(
        recorded_at=current.recorded_at,
        zone_count=current.zone_count,
        asset_count=current.asset_count,
        conduit_count=current.conduit_count,
        compliance_score=current.compliance_score,
        risk_score=current.risk_score,
        error_count=current.error_count,
        warning_count=current.warning_count,
    )

    # Calculate 7-day trends by comparing latest to value ~7 days ago
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    old_snapshots = [s for s in snapshots if s.recorded_at <= seven_days_ago]
    old_ref = old_snapshots[-1] if old_snapshots else snapshots[0]

    def _trend(current_val: float, old_val: float) -> TrendDirection:
        change = current_val - old_val
        threshold = 0.5  # Small threshold to avoid noise
        if abs(change) < threshold:
            direction = "stable"
        elif change > 0:
            direction = "up"
        else:
            direction = "down"
        return TrendDirection(value=current_val, direction=direction, change=round(change, 2))

    compliance_scores = [s.compliance_score for s in snapshots]
    risk_scores = [s.risk_score for s in snapshots]

    return AnalyticsSummary(
        current=current_point,
        compliance_trend=_trend(current.compliance_score, old_ref.compliance_score),
        risk_trend=_trend(current.risk_score, old_ref.risk_score),
        zone_count_trend=_trend(float(current.zone_count), float(old_ref.zone_count)),
        asset_count_trend=_trend(float(current.asset_count), float(old_ref.asset_count)),
        min_compliance=min(compliance_scores) if compliance_scores else None,
        max_compliance=max(compliance_scores) if compliance_scores else None,
        min_risk=min(risk_scores) if risk_scores else None,
        max_risk=max(risk_scores) if risk_scores else None,
        snapshot_count=len(snapshots),
    )
